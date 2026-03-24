"""
Volunteer Pass Generator v7 — Clean & Polished
================================================
Complete redesign. Neat, professional, print-ready A6.

Requirements: pip install reportlab openpyxl pillow

Usage:
    python volunteer_pass_v7.py --preview
    python volunteer_pass_v7.py --input volunteers.xlsx
    python volunteer_pass_v7.py --input volunteers.xlsx --batch-size 500
    python volunteer_pass_v7.py --template
"""

import argparse, io, math, os, sys
from PIL import Image, ImageDraw, ImageFont
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import openpyxl

# Allow large logo images
Image.MAX_IMAGE_PIXELS = None

# ── Fonts — loaded from a 'fonts' folder next to this script ──────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
FONT_DIR   = os.path.join(SCRIPT_DIR, 'fonts')

def font(name):
    return os.path.join(FONT_DIR, name)

pdfmetrics.registerFont(TTFont('PP-Bold',  font('Poppins-Bold.ttf')))
pdfmetrics.registerFont(TTFont('PP',       font('Poppins-Regular.ttf')))
pdfmetrics.registerFont(TTFont('PP-Light', font('Poppins-Light.ttf')))
pdfmetrics.registerFont(TTFont('PP-Med',   font('Poppins-Medium.ttf')))
DEVA_BOLD = font('NotoSansDevanagari-Bold.ttf')
DEVA_REG  = font('NotoSansDevanagari-Regular.ttf')
SS = 4  # supersampling scale

# ── A6 landscape: 148 × 105 mm ────────────────────────────────────
MM = 2.8346
CW = 148 * MM   # 419.5 pt
CH = 105 * MM   # 297.6 pt

# ── Palette ───────────────────────────────────────────────────────
MAROON  = colors.HexColor('#7B1C1C')
MAROON2 = colors.HexColor('#9B2222')
GOLD    = colors.HexColor('#B8922A')
GOLD2   = colors.HexColor('#D4A843')
WHITE   = colors.white
OFFWHITE= colors.HexColor('#FAFAFA')
INK     = colors.HexColor('#1A1A1A')
GREY    = colors.HexColor('#777777')
LGREY   = colors.HexColor('#F0F0F0')
DIVIDER = colors.HexColor('#E2E2E2')
CREAM   = colors.HexColor('#FDF8F0')

VERIFY_BASE_URL = "https://srjbtk-verify.caakshayshukla.workers.dev"

# ══════════════════════════════════════════════════════════════════
# QR ENCODER
# ══════════════════════════════════════════════════════════════════
class _QR:
    _L=[0]*256;_E=[0]*256;_rdy=False
    @classmethod
    def _init(cls):
        if cls._rdy:return
        x=1
        for i in range(255):cls._E[i]=x;cls._L[x]=i;x<<=1;x^=0x11D if x&0x100 else 0
        cls._E[255]=cls._E[0];cls._rdy=True
    @classmethod
    def _mul(cls,a,b):return 0 if(a==0 or b==0) else cls._E[(cls._L[a]+cls._L[b])%255]
    @classmethod
    def _poly(cls,n):
        q=[1]
        for j in range(n):
            t=[0]*(len(q)+1)
            for k,c in enumerate(q):t[k]^=c;t[k+1]^=cls._mul(c,cls._E[j])
            q=t
        return q
    @classmethod
    def _rs(cls,data,n):
        cls._init();gen=cls._poly(n);msg=list(data)+[0]*n
        for i in range(len(data)):
            cf=msg[i]
            if cf:
                for j in range(1,len(gen)):msg[i+j]^=cls._mul(gen[j],cf)
        return msg[len(data):]
    _VD={1:(21,19,7,1),2:(25,34,10,1),3:(29,55,15,1),4:(33,80,20,1),5:(37,108,26,1),
         6:(41,136,18,2),7:(45,156,20,2),8:(49,194,24,2),9:(53,232,30,2),10:(57,274,18,4)}
    _AL={2:[6,18],3:[6,22],4:[6,26],5:[6,30],6:[6,34],7:[6,22,38],
         8:[6,24,42],9:[6,26,46],10:[6,28,50]}
    def encode(self,text):
        data=text.encode('iso-8859-1','replace')
        ver=next((v for v in range(1,11) if len(data)<=self._VD[v][1]),None)
        if not ver:raise ValueError("Too long")
        sz,cap,en,bl=self._VD[ver];bits=[0,1,0,0]
        for i in range(7,-1,-1):bits.append((len(data)>>i)&1)
        for b in data:
            for i in range(7,-1,-1):bits.append((b>>i)&1)
        for _ in range(min(4,cap*8-len(bits))):bits.append(0)
        while len(bits)%8:bits.append(0)
        pi=0
        while len(bits)<cap*8:
            for i in range(7,-1,-1):bits.append(([0xEC,0x11][pi%2]>>i)&1);pi+=1
        cws=[sum(bits[i+j]<<(7-j) for j in range(8)) for i in range(0,len(bits),8)]
        bs=len(cws)//bl;dbs=[cws[b*bs:(b+1)*bs] for b in range(bl)]
        ebs=[self._rs(db,en) for db in dbs];fin=[]
        for i in range(max(len(b) for b in dbs)):
            for b in dbs:
                if i<len(b):fin.append(b[i])
        for i in range(max(len(b) for b in ebs)):
            for b in ebs:
                if i<len(b):fin.append(b[i])
        st=[]
        for byte in fin:
            for i in range(7,-1,-1):st.append((byte>>i)&1)
        st+=[0]*([0,7,7,7,7,7,0,0,0,0][ver-1])
        N=sz;mat=[[None]*N for _ in range(N)];fn=[[False]*N for _ in range(N)]
        def sf(r,c,v):
            if 0<=r<N and 0<=c<N:mat[r][c]=v;fn[r][c]=True
        def fp(tr,tc):
            for r in range(7):
                for c in range(7):sf(tr+r,tc+c,r in(0,6)or c in(0,6)or(2<=r<=4 and 2<=c<=4))
            for i in range(8):sf(tr-1,tc+i,False);sf(tr+i,tc-1,False);sf(tr+7,tc+i,False);sf(tr+i,tc+7,False)
        fp(0,0);fp(0,N-7);fp(N-7,0)
        if ver>=2:
            for r in self._AL.get(ver,[]):
                for c in self._AL.get(ver,[]):
                    if mat[r][c] is None:
                        for dr in range(-2,3):
                            for dc in range(-2,3):sf(r+dr,c+dc,abs(dr)==2 or abs(dc)==2 or(dr==0 and dc==0))
        for i in range(8,N-8):sf(6,i,i%2==0);sf(i,6,i%2==0)
        sf(N-8,8,True)
        for i in range(6):fn[8][i]=True;fn[i][8]=True
        fn[8][7]=True;fn[7][8]=True;fn[8][8]=True
        for i in range(N-8,N):fn[8][i]=True;fn[i][8]=True
        d=-1;row=N-1;col=N-1;idx=0
        while col>=1:
            if col==6:col-=1
            for _ in range(N):
                for dc in range(2):
                    c2=col-dc
                    if not fn[row][c2] and idx<len(st):
                        b=st[idx];b^=1 if(row+c2)%2==0 else 0;mat[row][c2]=bool(b);idx+=1
                row+=d
            d=-d;row+=d;col-=2
        fmt=0b01000;g=0b10100110111;tmp=fmt<<10
        for i in range(14,9,-1):
            if tmp&(1<<i):tmp^=g<<(i-10)
        fmt=(fmt<<10|tmp)^0b101010000010010
        fb=[(fmt>>i)&1 for i in range(14,-1,-1)]
        for i,p in enumerate(list(range(6))+[7,8]+list(range(N-7,N))):
            if i<15:mat[8][p]=bool(fb[i])
        for i,p in enumerate((list(range(6))+[7,8]+list(range(N-7,N)))[::-1][:8]+list(range(N-7,N))):
            if i<15:mat[p][8]=bool(fb[i])
        for r in range(N):
            for c in range(N):
                if mat[r][c] is None:mat[r][c]=False
        return mat

def qr_image(text, px=9):
    mat=_QR().encode(text[:300]);N=len(mat);Q=2;total=(N+2*Q)*px
    img=Image.new('RGB',(total,total),(255,255,255));draw=ImageDraw.Draw(img)
    for r in range(N):
        for c in range(N):
            if mat[r][c]:
                x0=(c+Q)*px;y0=(r+Q)*px
                draw.rectangle([x0,y0,x0+px-1,y0+px-1],fill=(20,20,20))
    return img


# ══════════════════════════════════════════════════════════════════
# DEVANAGARI — supersampled FreeSerif
# ══════════════════════════════════════════════════════════════════
def deva(text, pt=22, bold=False, color=(26,26,26)):
    """
    Render Devanagari using Windows GDI (ctypes) — zero extra dependencies.
    Windows Uniscribe shapes Devanagari conjuncts perfectly.
    Falls back to PIL on non-Windows systems.
    """
    import sys
    font_path = DEVA_BOLD if bold else DEVA_REG

    if sys.platform == "win32":
        try:
            import ctypes
            import ctypes.wintypes as wt
            import numpy as np

            gdi32 = ctypes.WinDLL("gdi32")
            user32 = ctypes.WinDLL("user32")

            # Register font file temporarily
            FR_PRIVATE = 0x10
            gdi32.AddFontResourceExW(font_path, FR_PRIVATE, None)

            # Get font family name from TTF
            from fontTools.ttLib import TTFont as FTFont
            ft = FTFont(font_path)
            family = ft["name"].getDebugName(1) or "NotoSansDevanagari"
            ft.close()

            dpi    = 384  # render at 4× then downscale → sharp crisp output
            h_px   = -int(pt * dpi / 72)
            weight = 700 if bold else 400

            hfont = gdi32.CreateFontW(
                h_px, 0, 0, 0, weight, 0, 0, 0,
                1, 0, 0, 5, 0, family  # 5=CLEARTYPE_QUALITY
            )

            hdc_screen = user32.GetDC(0)
            hdc = gdi32.CreateCompatibleDC(hdc_screen)
            old_font = gdi32.SelectObject(hdc, hfont)
            gdi32.SetBkMode(hdc, 1)  # TRANSPARENT

            # Measure text size
            class SIZE(ctypes.Structure):
                _fields_ = [("cx", ctypes.c_long), ("cy", ctypes.c_long)]
            sz = SIZE()
            gdi32.GetTextExtentPoint32W(hdc, text, len(text), ctypes.byref(sz))
            w = sz.cx + 30
            h = abs(h_px) * 3  # extra tall for Devanagari matras top/bottom

            # Create 32-bit DIB
            class BITMAPINFOHEADER(ctypes.Structure):
                _fields_ = [
                    ("biSize", wt.DWORD), ("biWidth", ctypes.c_long),
                    ("biHeight", ctypes.c_long), ("biPlanes", wt.WORD),
                    ("biBitCount", wt.WORD), ("biCompression", wt.DWORD),
                    ("biSizeImage", wt.DWORD), ("biXPelsPerMeter", ctypes.c_long),
                    ("biYPelsPerMeter", ctypes.c_long), ("biClrUsed", wt.DWORD),
                    ("biClrImportant", wt.DWORD),
                ]
            bmi = BITMAPINFOHEADER()
            bmi.biSize     = ctypes.sizeof(BITMAPINFOHEADER)
            bmi.biWidth    = w
            bmi.biHeight   = -h  # top-down
            bmi.biPlanes   = 1
            bmi.biBitCount = 32
            bmi.biCompression = 0

            bits = ctypes.c_void_p()
            hbmp = gdi32.CreateDIBSection(
                hdc, ctypes.byref(bmi), 0, ctypes.byref(bits), None, 0)
            gdi32.SelectObject(hdc, hbmp)

            r, g, b = color
            # Choose background opposite to text for maximum contrast
            luma_text = int(0.299*r + 0.587*g + 0.114*b)
            if luma_text > 128:
                # Light text — use black background
                gdi32.PatBlt(hdc, 0, 0, w, h, 0x00000042)  # BLACKNESS
                bg_luma = 0
            else:
                # Dark text — use white background
                gdi32.PatBlt(hdc, 0, 0, w, h, 0x00F00021)  # WHITENESS
                bg_luma = 255

            # Draw text
            gdi32.SetTextColor(hdc, (b << 16) | (g << 8) | r)  # GDI = BGR
            gdi32.TextOutW(hdc, 4, abs(h_px)//2, text, len(text))
            gdi32.GdiFlush()

            # Read bitmap pixels
            arr = np.frombuffer(
                (ctypes.c_uint8 * (w * h * 4)).from_address(bits.value),
                dtype=np.uint8
            ).reshape(h, w, 4).copy()

            # BGRA → RGBA, build alpha from contrast with background
            pixel_luma = (0.299*arr[:,:,2].astype(float) +
                          0.587*arr[:,:,1].astype(float) +
                          0.114*arr[:,:,0].astype(float))
            if bg_luma == 0:
                # Black bg: brighter pixels = more text
                alpha = np.clip(pixel_luma * 255 / max(luma_text, 1), 0, 255).astype(np.uint8)
            else:
                # White bg: darker pixels = more text
                alpha = np.clip((255 - pixel_luma) * 255 / max(255 - luma_text, 1), 0, 255).astype(np.uint8)
            rgba = np.zeros((h, w, 4), dtype=np.uint8)
            rgba[:,:,0] = r
            rgba[:,:,1] = g
            rgba[:,:,2] = b
            rgba[:,:,3] = alpha

            # Cleanup
            gdi32.SelectObject(hdc, old_font)
            gdi32.DeleteObject(hfont)
            gdi32.DeleteObject(hbmp)
            gdi32.DeleteDC(hdc)
            user32.ReleaseDC(0, hdc_screen)
            gdi32.RemoveFontResourceExW(font_path, FR_PRIVATE, None)

            img  = Image.fromarray(rgba, "RGBA")
            bbox = img.getbbox()
            if bbox:
                img = img.crop(bbox)
                out = Image.new("RGBA", (img.width+8, img.height+4), (0,0,0,0))
                out.paste(img, (4, 2))
                img = out
            # Downscale 4x for sharp antialiased output
            return img.resize((max(1,img.width//4), max(1,img.height//4)), Image.LANCZOS)

        except Exception:
            pass  # fall through to PIL

    # ── Linux: uharfbuzz + freetype-py shaping (proper conjuncts) ──
    try:
        import uharfbuzz as hb
        import freetype
        import numpy as np
        import logging
        log = logging.getLogger(__name__)

        SS = 4
        px = int(pt * SS * 1.33)

        # Load font into HarfBuzz
        with open(font_path, 'rb') as f:
            font_data = f.read()
        hb_blob = hb.Blob(font_data)
        hb_face = hb.Face(hb_blob)
        hb_font = hb.Font(hb_face)
        hb_font.scale = (px * 64, px * 64)

        # Shape the text
        buf = hb.Buffer()
        buf.add_str(text)
        buf.guess_segment_properties()
        hb.shape(hb_font, buf, {})

        infos  = buf.glyph_infos
        posits = buf.glyph_positions

        total_w = max(sum(p.x_advance for p in posits) // 64, 1)
        total_h = int(px * 1.8)  # generous height for matras
        pad = int(px * 0.2)

        # Render glyphs with freetype into numpy array
        face = freetype.Face(font_path)
        face.set_pixel_sizes(0, px)

        img_w = total_w + pad * 2
        img_h = total_h + pad * 2
        canvas = np.zeros((img_h, img_w), dtype=np.float32)
        baseline = int(px * 1.3) + pad
        pen_x = pad * 64

        for info, pos in zip(infos, posits):
            glyph_id = info.codepoint
            try:
                face.load_glyph(glyph_id, freetype.FT_LOAD_RENDER)
            except Exception:
                pen_x += pos.x_advance
                continue

            bmp = face.glyph.bitmap
            if bmp.rows == 0 or bmp.width == 0:
                pen_x += pos.x_advance
                continue

            bx = (pen_x + pos.x_offset) // 64 + face.glyph.bitmap_left
            by = baseline - face.glyph.bitmap_top + pos.y_offset // 64

            # Convert bitmap buffer to numpy
            bmp_arr = np.array(bmp.buffer, dtype=np.uint8).reshape(bmp.rows, bmp.pitch)
            bmp_arr = bmp_arr[:, :bmp.width].astype(np.float32)

            # Clip to canvas bounds
            x0 = max(bx, 0); x1 = min(bx + bmp.width, img_w)
            y0 = max(by, 0); y1 = min(by + bmp.rows, img_h)
            bx0 = x0 - bx; bx1 = bx0 + (x1 - x0)
            by0 = y0 - by; by1 = by0 + (y1 - y0)

            if x1 > x0 and y1 > y0:
                canvas[y0:y1, x0:x1] = np.minimum(
                    255.0,
                    canvas[y0:y1, x0:x1] + bmp_arr[by0:by1, bx0:bx1]
                )

            pen_x += pos.x_advance

        # Crop to actual content
        mask = canvas > 8
        if not mask.any():
            raise ValueError("Empty render")
        rows_with = np.where(mask.any(axis=1))[0]
        cols_with = np.where(mask.any(axis=0))[0]
        y0c, y1c = max(rows_with[0] - 2, 0), min(rows_with[-1] + 4, img_h)
        x0c, x1c = max(cols_with[0] - 2, 0), min(cols_with[-1] + 4, img_w)
        canvas = canvas[y0c:y1c, x0c:x1c]

        cr, cg, cb = color
        alpha = np.clip(canvas, 0, 255).astype(np.uint8)
        rgba  = np.zeros((*alpha.shape, 4), dtype=np.uint8)
        rgba[:, :, 0] = cr
        rgba[:, :, 1] = cg
        rgba[:, :, 2] = cb
        rgba[:, :, 3] = alpha

        hi = Image.fromarray(rgba, 'RGBA')
        # Downscale for crisp antialiased output
        out_w = max(1, hi.width  // SS)
        out_h = max(1, hi.height // SS)
        return hi.resize((out_w, out_h), Image.LANCZOS)

    except Exception as e:
        import logging
        logging.getLogger(__name__).warning(f"uharfbuzz render failed for '{text[:20]}': {e}")
        pass  # fall through to basic PIL

    # ── Basic PIL fallback (last resort, no shaping) ───────────────
    ss = 8 if pt < 14 else 4
    px = int(pt * ss * 1.33)
    try:   pil_font = ImageFont.truetype(font_path, px)
    except:return None
    probe = Image.new("RGBA",(1,1))
    bb    = ImageDraw.Draw(probe).textbbox((0,0), text, font=pil_font)
    pad   = 8
    w,h   = bb[2]-bb[0]+pad*2, bb[3]-bb[1]+pad*2
    hi    = Image.new("RGBA",(w,h),(0,0,0,0))
    ImageDraw.Draw(hi).text((-bb[0]+pad,-bb[1]+pad), text,
                            font=pil_font, fill=color+(255,))
    return hi.resize((max(1,w//ss), max(1,h//ss)), Image.LANCZOS)


def place_deva(c, text, x, y, pt=22, bold=False, color=(26,26,26), max_w=None):
    img = deva(text, pt=pt, bold=bold, color=color)
    if img is None: return 0, 0
    w_pt = img.width*0.75; h_pt = img.height*0.75
    if max_w and w_pt > max_w:
        h_pt *= max_w/w_pt; w_pt = max_w
    buf=io.BytesIO(); img.save(buf,'PNG'); buf.seek(0)
    c.drawImage(ImageReader(buf), x, y, w_pt, h_pt, mask='auto')
    return w_pt, h_pt

def irl(img):
    buf=io.BytesIO(); img.save(buf,'PNG'); buf.seek(0)
    return ImageReader(buf)


# ══════════════════════════════════════════════════════════════════
# PASS DRAWING — v7 clean design
# ══════════════════════════════════════════════════════════════════
def draw_pass(c, vol):
    vol_id  = str(vol.get('id',  'VOL00001'))
    name_en = str(vol.get('name','Volunteer Name'))
    name_hi = str(vol.get('name_hi',''))
    role    = str(vol.get('role',''))
    aadhaar = mask_aadhaar(str(vol.get('aadhaar','')))
    mobile  = str(vol.get('mobile',''))
    org     = str(vol.get('org',  'श्री राम जन्मभूमि तीर्थ क्षेत्र'))
    perm      = str(vol.get('permission', '')).strip()
    pass_type = str(vol.get('pass_type',   'standard')).strip().lower()
    expiry    = str(vol.get('expiry',      '')).strip()

    # Colour theme based on pass type
    THEMES = {
        'standard': {'primary': colors.HexColor('#7B1C1C'), 'accent': colors.HexColor('#B8922A'), 'label': ''},
        'vip':      {'primary': colors.HexColor('#1A1A2E'), 'accent': colors.HexColor('#FFD700'), 'label': 'VIP'},
        'media':    {'primary': colors.HexColor('#0D47A1'), 'accent': colors.HexColor('#90CAF9'), 'label': 'MEDIA'},
        'security': {'primary': colors.HexColor('#1B5E20'), 'accent': colors.HexColor('#A5D6A7'), 'label': 'SECURITY'},
        'medical':  {'primary': colors.HexColor('#880E4F'), 'accent': colors.HexColor('#F48FB1'), 'label': 'MEDICAL'},
    }
    theme = THEMES.get(pass_type, THEMES['standard'])
    T_PRIMARY = theme['primary']
    T_ACCENT  = theme['accent']
    T_LABEL   = theme['label']
    event   = str(vol.get('event_label','वर्ष प्रतिपदा समारोह - कार्यकर्ता पास - 20 मार्च तक मान्य'))

    # ── White base ───────────────────────────────────────────────
    c.setFillColor(WHITE)
    c.rect(0, 0, CW, CH, fill=1, stroke=0)

    # ══ TOP HEADER BAND ══════════════════════════════════════════
    HDR = 23*MM   # balanced header height
    c.setFillColor(T_PRIMARY)
    c.rect(0, CH-HDR, CW, HDR, fill=1, stroke=0)

    # Gold bottom edge of header
    c.setFillColor(T_ACCENT)
    c.rect(0, CH-HDR-1*MM, CW, 1*MM, fill=1, stroke=0)

    # Logo — optional, fitted neatly inside header
    LOGO_PATH = os.path.join(SCRIPT_DIR, 'srjbtk_logo_official.png')
    text_x = 4*MM
    if os.path.exists(LOGO_PATH):
        logo   = Image.open(LOGO_PATH).convert('RGBA')
        logo_h = HDR - 5*MM
        logo_w = logo_h * logo.width / logo.height
        logo   = logo.resize((int(logo_w*3), int(logo_h*3)), Image.LANCZOS)
        logo_y = CH - HDR + 2.5*MM
        c.drawImage(irl(logo), 2*MM, logo_y, logo_w, logo_h, mask='auto')
        text_x = 2*MM + logo_w + 3*MM

    # Org + event — render on maroon swatch so text is always visible
    avail_px = int((CW - text_x - 3*MM) * 4)   # available width in pixels @4x
    hdr_h_px = int(HDR * 4)                      # header height in pixels @4x

    # Header text — use same deva() GDI renderer as names (perfect shaping)
    avail_pt = CW - text_x - 3*MM
    place_deva(c, org,   text_x, CH-HDR+HDR*0.48, pt=16, bold=True,
               color=(255,255,255), max_w=avail_pt)
    place_deva(c, event, text_x, CH-HDR+HDR*0.12, pt=11, bold=False,
               color=(255,224,144), max_w=avail_pt)

    # ══ RIGHT COLUMN — ID box + QR ═══════════════════════════════
    RC_W  = 32*MM                           # right column width
    RC_X  = CW - RC_W - 3*MM               # right column left edge
    PAD_V = 4*MM                            # vertical padding from header/bottom

    BODY_TOP = CH - HDR - 0.8*MM           # top of body area
    BODY_H   = BODY_TOP - PAD_V            # usable body height

    # ── ID box — top of right column ─────────────────────────────
    ID_H  = 13*MM
    ID_Y  = BODY_TOP - ID_H - 2*MM        # just below header gold line

    c.setFillColor(LGREY)
    c.roundRect(RC_X, ID_Y, RC_W, ID_H, 1.5*MM, fill=1, stroke=0)

    c.setFillColor(GREY); c.setFont('PP-Light', 5)
    c.drawCentredString(RC_X + RC_W/2, ID_Y + ID_H - 4.5*MM, 'VOLUNTEER ID')

    c.setFillColor(MAROON); c.setFont('PP-Bold', 8.5)
    c.drawCentredString(RC_X + RC_W/2, ID_Y + 2.5*MM, vol_id)

    # ── QR code — directly below ID box ──────────────────────────
    QR_SIZE = RC_W
    QR_Y    = ID_Y - QR_SIZE - 3*MM       # 3mm gap below ID box

    if VERIFY_BASE_URL:
        qr_data = f"{VERIFY_BASE_URL}?id={vol_id}"
    else:
        qr_data = f"{vol_id}|{name_en}|{role}|ADH:{str(vol.get('aadhaar',''))[-4:]}"

    qr = qr_image(qr_data, px=9)
    qr = qr.resize((int(QR_SIZE*3), int(QR_SIZE*3)), Image.NEAREST)
    c.drawImage(irl(qr), RC_X, QR_Y, QR_SIZE, QR_SIZE)

    c.setFillColor(GREY); c.setFont('PP-Light', 4.5)
    c.drawCentredString(RC_X + RC_W/2, QR_Y - 2.5*MM, 'SCAN TO VERIFY')

    # Vertical separator line between left and right
    c.setStrokeColor(DIVIDER); c.setLineWidth(0.4)
    c.line(RC_X - 3*MM, BODY_TOP - 2*MM, RC_X - 3*MM, PAD_V + 2*MM)

    # ══ LEFT COLUMN — Name + Fields ══════════════════════════════
    LC_X  = 4*MM
    LC_W  = RC_X - 3*MM - LC_X - 2*MM
    cur_y = BODY_TOP - 3*MM

    # Hindi name
    if name_hi:
        _, h_pt = place_deva(c, name_hi, LC_X, cur_y - 9.5*MM,
                             pt=24, bold=True, color=(26,26,26), max_w=LC_W)
        cur_y -= 10.5*MM

    # English name
    cur_y -= 7.5*MM
    c.setFillColor(INK); c.setFont('PP-Bold', 11)
    nm = name_en
    while c.stringWidth(nm,'PP-Bold',11) > LC_W and len(nm)>3: nm=nm[:-1]
    c.drawString(LC_X, cur_y, nm)

    # Role pill removed

    # Pass type label (if not standard)
    if T_LABEL:
        lw = c.stringWidth(T_LABEL, 'PP-Bold', 7) + 6*MM
        c.setFillColor(T_PRIMARY)
        c.roundRect(LC_X, cur_y - 5.5*MM, lw, 5*MM, radius=2*MM, fill=1, stroke=0)
        c.setFillColor(colors.white)
        c.setFont('PP-Bold', 7)
        c.drawString(LC_X + 3*MM, cur_y - 3.5*MM, T_LABEL)
        cur_y -= 7*MM

    # Divider
    cur_y -= 5*MM
    c.setStrokeColor(DIVIDER); c.setLineWidth(0.5)
    c.line(LC_X, cur_y, LC_X + LC_W, cur_y)
    cur_y -= 4*MM

    # Permission pill — shows exact permission granted
    if perm:
        perm_lower = perm.lower()
        if 'no ' in perm_lower or 'not allowed' in perm_lower or 'restricted' in perm_lower:
            pill_bg = colors.HexColor('#B71C1C')  # red — nothing allowed
        else:
            pill_bg = colors.HexColor('#1B5E20')  # green — items allowed

        PILL_H  = 7*MM
        PILL_R  = 2.5*MM  # corner radius
        PILL_PAD = 3*MM

        # Pill background
        c.setFillColor(pill_bg)
        c.roundRect(LC_X, cur_y - PILL_H, LC_W, PILL_H,
                    radius=PILL_R, fill=1, stroke=0)

        # Permission label — small caps above
        c.setFillColor(colors.HexColor('#FFFFFF'))
        c.setFont('PP-Bold', 5)
        c.drawString(LC_X + PILL_PAD, cur_y - 2.0*MM, 'PERMISSION GRANTED')

        # Permission text — larger, clear, with spacing
        perm_size = 7
        while c.stringWidth(perm, 'PP-Bold', perm_size) > LC_W - PILL_PAD*2 and perm_size > 5:
            perm_size -= 0.3
        c.setFont('PP-Bold', perm_size)
        c.drawString(LC_X + PILL_PAD, cur_y - PILL_H + 2.2*MM, perm)

        cur_y -= PILL_H + 2*MM

    # Fields
    def field(hi_lbl, en_lbl, val):
        nonlocal cur_y
        if not str(val).strip(): return

        # Hindi label — white swatch background so always visible
        SS = 8  # higher supersampling for crisp small text
        px_size = int(9 * SS * 1.33)
        lbl_font = None
        for fp in [DEVA_BOLD,
                   os.path.join(FONT_DIR, 'FreeSerifBold.ttf'),
                   '/usr/share/fonts/truetype/freefont/FreeSerifBold.ttf']:
            try:
                lbl_font = ImageFont.truetype(fp, px_size)
                break
            except: continue

        if lbl_font:
            probe = Image.new('RGBA', (1,1))
            bb    = ImageDraw.Draw(probe).textbbox((0,0), hi_lbl, font=lbl_font)
            pad   = 4
            w, h  = bb[2]-bb[0]+pad*2, bb[3]-bb[1]+pad*2
            swatch = Image.new('RGBA', (w, h), (255,255,255,255))
            ImageDraw.Draw(swatch).text((-bb[0]+pad,-bb[1]+pad), hi_lbl,
                                        font=lbl_font, fill=(80,20,20,255))
            out = swatch.resize((max(1,w//SS), max(1,h//SS)), Image.LANCZOS)
            ow = out.width * 0.75 * 0.5; oh = out.height * 0.75 * 0.5  # compensate for SS=8
            buf = io.BytesIO(); out.save(buf,'PNG'); buf.seek(0)
            c.drawImage(ImageReader(buf), LC_X, cur_y - 4.5*MM, ow, oh)

        # English label
        c.setFillColor(MAROON); c.setFont('PP-Med', 6)
        c.drawString(LC_X + LC_W*0.43, cur_y - 2.8*MM, f'/ {en_lbl}')
        # Value
        c.setFillColor(INK); c.setFont('PP-Bold', 9)
        c.drawString(LC_X, cur_y - 8*MM, str(val)[:30])
        cur_y -= 12.5*MM

    field('आधार',   'AADHAAR', aadhaar)
    field('मोबाइल', 'MOBILE',  mobile)

    # ══ BOTTOM STRIP ══════════════════════════════════════════════
    c.setFillColor(T_PRIMARY)
    c.rect(0, 0, CW, 5.5*MM, fill=1, stroke=0)
    c.setFillColor(GOLD)
    c.rect(0, 5.5*MM, CW, 0.5*MM, fill=1, stroke=0)
    c.setFillColor(colors.HexColor('#CCCCCC')); c.setFont('PP-Light', 4.5)
    c.drawCentredString(CW/2, 1.8*MM,
        'This pass is non-transferable  ·  Valid only with original photo ID')

    # ══ OUTER BORDER ══════════════════════════════════════════════
    c.setStrokeColor(MAROON); c.setLineWidth(1.2)
    c.rect(0.6*MM, 0.6*MM, CW-1.2*MM, CH-1.2*MM, fill=0, stroke=1)
    c.setStrokeColor(GOLD); c.setLineWidth(0.35)
    c.rect(1.5*MM, 1.5*MM, CW-3*MM, CH-3*MM, fill=0, stroke=1)


def mask_aadhaar(v):
    d=''.join(filter(str.isdigit,v))
    return f'XXXX XXXX {d[-4:]}' if len(d)>=4 else 'XXXX XXXX XXXX'


# ══════════════════════════════════════════════════════════════════
# EXCEL
# ══════════════════════════════════════════════════════════════════
COLS = {
    'name':        ['name','full name','volunteer name','name (english)'],
    'name_hi':     ['name_hi','hindi name','naam','नाम','name (hindi)','hindi','name(hindi)'],
    'role':        ['role','title','designation'],
    'id':          ['id','id number','volunteer id','kramank','id number'],
    'aadhaar':     ['aadhaar','aadhar','uid'],
    'mobile':      ['mobile','phone','contact'],
    'org':         ['org','organization','organisation'],
    'event_label': ['event','event label','event name'],
    'permission':  ['permission','permissions','access','अनुमति'],
    'pass_type':   ['pass type','pass_type','type','pass'],
    'expiry':      ['expiry','expiry date','valid until','validity','valid till'],
}

def read_excel(path):
    wb=openpyxl.load_workbook(path,data_only=True)
    ws=wb.active; rows=list(ws.iter_rows(values_only=True))
    if not rows: raise ValueError("Empty")
    header=[str(h).strip().lower() if h else '' for h in rows[0]]
    cm={}
    for f,aliases in COLS.items():
        for i,h in enumerate(header):
            if h in aliases: cm[f]=i; break
    vols=[]
    for ri,row in enumerate(rows[1:],2):
        if not any(row): continue
        v={f:(row[i] if i<len(row) and row[i] is not None else '') for f,i in cm.items()}
        if not v.get('id'): v['id']=f'VOL{ri:05d}'
        # Normalise expiry date to DD-MM-YYYY string
        if v.get('expiry'):
            import datetime as dt
            exp = v['expiry']
            if isinstance(exp, (dt.date, dt.datetime)):
                v['expiry'] = exp.strftime('%d-%m-%Y')
            else:
                v['expiry'] = str(exp).strip()
        vols.append(v)
    return vols

def sample_data():
    ORG='श्री राम जन्मभूमि तीर्थ क्षेत्र'
    EVT='वर्ष प्रतिपदा समारोह - कार्यकर्ता पास - 20 मार्च तक मान्य'
    return [
        {'name':'Ashwani Tandon',  'name_hi':'अश्वनी टंडन',  'role':'Medical Camp', 'permission':'Mobile Phone Allowed', 'pass_type':'standard', 'expiry':'2026-03-20',
         'id':'VLTR0379','aadhaar':'9876543219163','mobile':'9246149248','org':ORG,'event_label':EVT},
        {'name':'Priya Sharma',    'name_hi':'प्रिया शर्मा',  'role':'Registration Desk', 'permission':'Mobile + Video Camera + Laptop Allowed', 'pass_type':'media', 'expiry':'2026-03-20',
         'id':'VLTR0380','aadhaar':'111122223333', 'mobile':'9876543210','org':ORG,'event_label':EVT},
        {'name':'Arjun Mehta',     'name_hi':'अर्जुन मेहता',  'role':'Security', 'permission':'No Items Allowed', 'pass_type':'vip', 'expiry':'2026-03-20',
         'id':'VLTR0381','aadhaar':'444455556666', 'mobile':'9123456780','org':ORG,'event_label':EVT},
        {'name':'Kavitha Nair',    'name_hi':'कविता नायर',    'role':'Security', 'pass_type':'security',
         'permission':'No Mobile', 'expiry':'20-03-2026',
         'id':'VLTR0382','aadhaar':'777788889999', 'mobile':'9988776655','org':ORG,'event_label':EVT},
        {'name':'Sunita Reddy',    'name_hi':'सुनीता रेड्डी', 'role':'Medical Team', 'pass_type':'medical',
         'permission':'Mobile Only', 'expiry':'20-03-2026',
         'id':'VLTR0383','aadhaar':'123412341234', 'mobile':'9001234567','org':ORG,'event_label':EVT},
    ]


# ══════════════════════════════════════════════════════════════════
# PDF WRITER
# ══════════════════════════════════════════════════════════════════
def generate(vols, out, batch_size=None, progress=True):
    if batch_size:
        base,ext=os.path.splitext(out); n=math.ceil(len(vols)/batch_size)
        for b in range(n):
            chunk=vols[b*batch_size:(b+1)*batch_size]; p=f'{base}_batch{b+1:03d}{ext}'
            _write(chunk,p,progress,f'Batch {b+1}/{n}: ')
            print(f'  → {p}  ({len(chunk)} passes)')
    else:
        _write(vols,out,progress); print(f'\n✓ {out}  ({len(vols)} passes)')

def _write(vols, path, progress=True, prefix=''):
    c=canvas.Canvas(path,pagesize=(CW,CH)); total=len(vols)
    for i,vol in enumerate(vols):
        if progress and (i%25==0 or i==total-1):
            pct=(i+1)/total*100; bar='█'*int(pct//5)+'░'*(20-int(pct//5))
            print(f'\r  {prefix}[{bar}] {i+1}/{total} ({pct:.0f}%)',end='',flush=True)
        draw_pass(c,vol); c.showPage()
    c.save()
    if progress: print()


# ══════════════════════════════════════════════════════════════════
# TEMPLATE
# ══════════════════════════════════════════════════════════════════
def create_template(out='volunteer_template_v7.xlsx'):
    from openpyxl.styles import Font, PatternFill, Alignment
    wb=openpyxl.Workbook(); ws=wb.active; ws.title='Volunteers'
    headers=['Name','Name (Hindi)','Role','ID Number','Aadhaar','Mobile','Organization','Event Label']
    widths=[22,22,18,12,16,14,32,32]
    for ci,(h,w) in enumerate(zip(headers,widths),1):
        cell=ws.cell(1,ci,h)
        cell.font=Font(bold=True,color='FFFFFF',size=10)
        cell.fill=PatternFill('solid',fgColor='7B1C1C')
        cell.alignment=Alignment(horizontal='center',vertical='center')
        ws.column_dimensions[cell.column_letter].width=w
    ws.row_dimensions[1].height=24
    for r,s in enumerate(sample_data(),2):
        for ci,k in enumerate(['name','name_hi','role','id','aadhaar','mobile','org','event_label'],1):
            ws.cell(r,ci,s.get(k,'')).font=Font(size=10)
    ws2=wb.create_sheet('Instructions')
    for r,l in enumerate([
        'VOLUNTEER PASS GENERATOR v7','',
        '1. Fill the Volunteers sheet with your data',
        '2. Keep srjbtk_logo_official.png in the same folder as the script',
        '3. Run: python volunteer_pass_v7.py --preview',
        '4. Run: python volunteer_pass_v7.py --input volunteers.xlsx --batch-size 500',
    ],1):
        ws2[f'A{r}']=l; ws2[f'A{r}'].font=Font(bold=(r==1),size=10)
    ws2.column_dimensions['A'].width=60
    wb.save(out); print(f'✓ Template: {out}')


# ══════════════════════════════════════════════════════════════════
# CLI
# ══════════════════════════════════════════════════════════════════
def main():
    p=argparse.ArgumentParser()
    p.add_argument('--input',      default='volunteers.xlsx')
    p.add_argument('--output',     default='volunteer_passes_v7.pdf')
    p.add_argument('--batch-size', type=int, default=None)
    p.add_argument('--preview',    action='store_true')
    p.add_argument('--template',   action='store_true')
    p.add_argument('--ids',        default=None,
                   help='Comma-separated IDs: VLTR0001,VLTR0005')
    p.add_argument('--name',       default=None,
                   help='Filter by name (partial match)')
    p.add_argument('--role',       default=None,
                   help='Filter by role (partial match)')
    p.add_argument('--from-row',   type=int, default=None,
                   help='Start from this row number in Excel (1=first data row)')
    p.add_argument('--to-row',     type=int, default=None,
                   help='End at this row number in Excel (inclusive)')
    args=p.parse_args()
    print('\u2554'+'\u2550'*38+'\u2557')
    print('\u2551  VOLUNTEER PASS GENERATOR  v7        \u2551')
    print('\u255a'+'\u2550'*38+'\u255d\n')
    if args.template: create_template(); return
    if args.preview:
        print('\u25ba Generating 5-pass preview...')
        generate(sample_data(),'volunteer_passes_v7_PREVIEW.pdf'); return
    if not os.path.exists(args.input):
        print(f'\u2717 Not found: {args.input}'); sys.exit(1)
    vols = read_excel(args.input)
    print(f'\u25ba {len(vols)} volunteers loaded')

    # Row range filter (1-based, refers to data rows excluding header)
    if args.from_row is not None or args.to_row is not None:
        fr = (args.from_row or 1) - 1        # convert to 0-based
        to = (args.to_row   or len(vols))     # inclusive
        vols = vols[fr:to]
        print(f'\u25ba Row range: {args.from_row or 1} to {args.to_row or len(vols)+fr} → {len(vols)} volunteers')

    # ID filter
    if args.ids:
        id_list = [i.strip().upper() for i in args.ids.split(',')]
        vols = [v for v in vols if str(v.get('id','')).upper() in id_list]
        print(f'\u25ba Filtered to {len(vols)} volunteers by ID')

    # Name filter
    if args.name:
        q = args.name.lower()
        vols = [v for v in vols if q in str(v.get('name','')).lower()
                                or q in str(v.get('name_hi','')).lower()]
        print(f'\u25ba Filtered to {len(vols)} volunteers by name')

    # Role filter
    if args.role:
        q = args.role.lower()
        vols = [v for v in vols if q in str(v.get('role','')).lower()]
        print(f'\u25ba Filtered to {len(vols)} volunteers by role')

    if not vols:
        print('\u2717 No matching volunteers found'); sys.exit(1)

    generate(vols, args.output, args.batch_size)

if __name__=='__main__':
    main()
